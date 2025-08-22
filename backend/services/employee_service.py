from sqlalchemy.orm import Session
from typing import List
import pandas as pd
from io import BytesIO

from models import employee as employee_model
from schemas import employee as employee_schema
from utils.security import get_password_hash, verify_password

def authenticate_user(db: Session, username: str, password: str):
    db_employee = db.query(employee_model.Employee).filter(employee_model.Employee.employee_no == username).first()
    if not db_employee:
        return False
    if not verify_password(password, db_employee.password):
        return False
    return db_employee

def create_employee(db: Session, employee: employee_schema.EmployeeCreate):
    hashed_password = get_password_hash(employee.password)
    db_employee = employee_model.Employee(
        **employee.dict(exclude={"password"}), 
        password=hashed_password
    )
    db.add(db_employee)
    db.commit()
    db.refresh(db_employee)
    return db_employee

def get_employees(db: Session, skip: int = 0, limit: int = 100, name: str = None):
    query = db.query(employee_model.Employee)
    if name:
        query = query.filter(employee_model.Employee.name.contains(name))
    employees = query.offset(skip).limit(limit).all()
    return employees

def get_employee(db: Session, employee_id: int):
    return db.query(employee_model.Employee).filter(employee_model.Employee.employee_id == employee_id).first()

def update_employee(db: Session, employee_id: int, employee: employee_schema.EmployeeCreate):
    db_employee = get_employee(db, employee_id)
    if db_employee:
        update_data = employee.dict(exclude_unset=True)
        if "password" in update_data:
            hashed_password = get_password_hash(update_data["password"])
            db_employee.password = hashed_password
        for key, value in update_data.items():
            if key != "password":
                setattr(db_employee, key, value)
        db.commit()
        db.refresh(db_employee)
    return db_employee

def delete_employee(db: Session, employee_id: int):
    db_employee = get_employee(db, employee_id)
    if db_employee:
        db.delete(db_employee)
        db.commit()
    return db_employee

def get_employee_by_employee_no(db: Session, employee_no: str):
    return db.query(employee_model.Employee).filter(employee_model.Employee.employee_no == employee_no).first()

def import_employees_from_file(db: Session, contents: bytes):
    try:
        df = pd.read_excel(BytesIO(contents))
        
        # 中文列名映射到英文字段名
        column_mapping = {
            '员工ID': 'employee_id',
            '员工编号': 'employee_no', 
            '姓名': 'name',
            '性别': 'gender',
            '电话': 'phone',
            '邮箱': 'email',
            '职位': 'position',
            '入职日期': 'hire_date',
            '是否管理员': 'is_admin',
            '密码': 'password'
        }
        
        # 重命名列
        df = df.rename(columns=column_mapping)
        
        imported_count = 0
        for index, row in df.iterrows():
            try:
                # 处理是否管理员字段
                if 'is_admin' in row:
                    row['is_admin'] = row['is_admin'] in ['是', True, 1, '1']
                
                # 如果没有密码，设置默认密码
                if 'password' not in row or pd.isna(row['password']):
                    row['password'] = '123456'  # 默认密码
                
                # 确保字符串字段的类型转换
                if 'phone' in row and not pd.isna(row['phone']):
                    row['phone'] = str(row['phone'])
                if 'employee_no' in row and not pd.isna(row['employee_no']):
                    row['employee_no'] = str(row['employee_no'])
                
                # 跳过employee_id字段（自动生成）
                row_dict = row.to_dict()
                if 'employee_id' in row_dict:
                    del row_dict['employee_id']
                
                employee_data = employee_schema.EmployeeCreate(**row_dict)
                
                # 检查员工编号是否已存在
                existing_employee = get_employee_by_employee_no(db, employee_data.employee_no)
                if existing_employee:
                    print(f"Employee with employee_no {employee_data.employee_no} already exists, skipping")
                    continue
                
                # 对密码进行哈希处理
                hashed_password = get_password_hash(employee_data.password)
                db_employee = employee_model.Employee(
                    **employee_data.dict(exclude={"password"}),
                    password=hashed_password
                )
                db.add(db_employee)
                imported_count += 1
                
            except Exception as e:
                print(f"Error importing row {index}: {str(e)}")
                continue
        
        db.commit()
        return imported_count
        
    except Exception as e:
        print(f"Error in import_employees_from_file: {str(e)}")
        db.rollback()
        raise e

def export_employees_to_excel(db: Session):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        employees = db.query(employee_model.Employee).all()
        print(f"Found {len(employees)} employees in database")
        
        if not employees:
            return None
        
        # 使用openpyxl直接创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "员工信息"
        
        # 设置表头
        headers = ['员工ID', '员工编号', '姓名', '性别', '电话', '邮箱', '职位', '入职日期', '是否管理员']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 填充数据
        for row, employee in enumerate(employees, 2):
            ws.cell(row=row, column=1, value=employee.employee_id)
            ws.cell(row=row, column=2, value=employee.employee_no)
            ws.cell(row=row, column=3, value=employee.name)
            ws.cell(row=row, column=4, value=employee.gender)
            ws.cell(row=row, column=5, value=employee.phone)
            ws.cell(row=row, column=6, value=employee.email)
            ws.cell(row=row, column=7, value=employee.position)
            ws.cell(row=row, column=8, value=employee.hire_date.strftime('%Y-%m-%d') if employee.hire_date else '')
            ws.cell(row=row, column=9, value='是' if employee.is_admin else '否')
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        print(f"Excel file created successfully with {len(employees)} records")
        return output
        
    except Exception as e:
        print(f"Error in export_employees_to_excel: {str(e)}")
        import traceback
        traceback.print_exc()
        raise e

def update_employee_password(db: Session, employee_id: int, new_password: str):
    """更新员工密码"""
    db_employee = get_employee(db, employee_id)
    if db_employee:
        hashed_password = get_password_hash(new_password)
        db_employee.password = hashed_password
        db.commit()
        db.refresh(db_employee)
    return db_employee